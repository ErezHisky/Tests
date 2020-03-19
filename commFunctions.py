import openpyxl
from Drivers.DigAtt64bit import *
from Drivers.StepEngineCtrl import *
from Drivers.SigGenCtrl import *
from Drivers.Terminal.SendMsg import *
from Drivers.Terminal.SendDataMsg import *


def ipusAll():
    DigAtt_putMax()
    signal_off()
    # Engine_GoToHome()
    # Hub_Gold_Unit_off()

def DigAtt_putMax():
    myAtt = DigAtt64()
    myAtt.set_attenuation(58.0)

def signal_off():
    mySig = MySignalGenerator()
    mySig.turnModOff()
    mySig.turnRFOff()

def Engine_GoToHome():
    my_eng = MyStepEngine("azimuth")
    my_eng.check_communication()
    my_eng.get_to_home()
    time.sleep(1)
    my_eng = MyStepEngine("elevation")
    my_eng.get_to_home()
    my_eng.close_engine()

def Hub_Gold_Unit_off():
    pass

def getHRprops(unitNumber):
    wb = openpyxl.load_workbook("D:/ExcelFiles/RH_prop_example.xlsx")
    ws = wb.active
    for val in range(2, ws.max_row + 1):
        if ws['A' + str(val)].value == unitNumber:
            return [ws['B'+str(val)].value, ws['C'+str(val)].value]

def CheckPingAllAngles(host, n=0):
    from multiping import MultiPing

    if (n > 0):
        avg = 0
        for i in range(n):
            avg += CheckPingAllAngles(host)
        avg = avg / n
    # Create a MultiPing object to test hosts / addresses
    mp = MultiPing([host])

    # Send the pings to those addresses
    mp.send()

    # With a 1 second timout, wait for responses (may return sooner if all
    # results are received).
    responses, no_responses = mp.receive(1)

    for addr, rtt in responses.items():
        RTT = rtt

    if no_responses:
        # Sending pings once more, but just to those addresses that have not
        # responded, yet.
        mp.send()
        responses, no_responses = mp.receive(1)
        RTT = -1

    return RTT


